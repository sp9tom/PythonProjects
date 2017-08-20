# coding=utf-8
__author__ = 'Tomasz'

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
import re
import openpyxl
import datetime

class Formularz():
    driver = webdriver.Firefox()
    lgn = "biuro@mila.org.pl"
    passwd = "mila121110"

    def __init__(self):
        self.driver.get("http://formularze.fed.org.pl/")
        self.driver.maximize_window()
        assert "Fundacja Edukacja dla Demokracji" in self.driver.title
        login = self.driver.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/div/table[2]/tbody/tr/td[3]/form/p[1]/input")
        login.send_keys(self.lgn)
        haslo = self.driver.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/div/table[2]/tbody/tr/td[3]/form/p[2]/input")
        haslo.send_keys(self.passwd)
        klZaloguj = self.driver.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/div/table[2]/tbody/tr/td[3]/form/p[3]/input")
        klZaloguj.click()
        sleep(2)
        klSprawozdanie = self.driver.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/div/table[3]/tbody/tr[2]/td[4]/a")
        klSprawozdanie.click()
        sleep(8)
        spr = self.driver.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/div/table/tbody/tr/td/div/form/table[2]/tbody/tr/td/table/tbody/tr/td/ul/li[2]/a")
        spr.click()
        sleep(2)
        assert "GENERATOR" in self.driver.title

    def wyloguj(self):
        klPowrot = self.driver.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/div/table/tbody/tr/td/div/form/div[1]/table/tbody/tr/td[1]/div")
        klPowrot.click()
        alert = self.driver.switch_to_alert()
        alert.accept()
        sleep(5)
        klWyloguj = self.driver.find_element_by_xpath("/html/body/table/tbody/tr/td/div/table[1]/tbody/tr/td/div[1]/a")
        klWyloguj.click()
        sleep(5)
        self.driver.close()

    def dodaj(self, input):
        if len(input):
            klDodajPozycje = self.driver.find_element_by_xpath(
                "/html/body/table/tbody/tr/td/div/table/tbody/tr/td/div/form/table[2]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/div[2]/span/table/tbody/tr[4]/td/div[2]/a")
            klDodajPozycje.click()
            window = self.driver.switch_to_active_element()

            poz1 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[1]/td[2]/input")
            sleep(1)
            poz1.click()
            poz1.send_keys(input[0])
            sleep(1)

            poz2 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[2]/td[2]/input")
            sleep(1)
            poz2.click()
            poz2.send_keys(input[1])
            sleep(1)

            poz3 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[3]/td[2]/input")
            sleep(1)
            poz3.click()
            poz3.send_keys(input[2])
            sleep(1)
            poz3.send_keys(Keys.ENTER)
            sleep(1)

            poz4 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[4]/td[2]/input")
            sleep(1)
            poz4.click()
            poz4.send_keys(input[3])
            sleep(1)
            poz4.send_keys(Keys.ENTER)
            sleep(1)

            window.find_element_by_xpath("//select[@class='input-sm']/option[text()='%s']" % input[4]).click() # poz5
            sleep(1)

            poz6 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[6]/td[2]/input")
            sleep(1)
            poz6.click()
            poz6.send_keys(input[5])
            sleep(1)

            poz7 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[7]/td[2]/input")
            sleep(1)
            poz7.click()
            poz7.send_keys(input[6])
            sleep(1)

            poz8 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[8]/td[2]/input")
            sleep(1)
            poz8.click()
            poz8.send_keys(input[7])
            sleep(1)

            poz9 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[9]/td[2]/input")
            sleep(1)
            poz9.click()
            poz9.send_keys(input[8])
            sleep(1)

            poz10 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[10]/td[2]/input")
            sleep(1)
            poz10.click()
            poz10.send_keys(input[9])
            sleep(1)

            poz11 = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/table/tbody/tr[11]/td[2]/input")
            sleep(1)
            poz11.click()
            poz11.send_keys(input[10])
            sleep(1)

            # klPorzuc = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/footer/p/button[2]")
            # klPorzuc.click()
            # sleep(2)

            klDodaj = window.find_element_by_xpath("/html/body/table/tbody/tr/td/div/div[1]/div[2]/div/footer/p/button[1]")
            klDodaj.click()
            sleep(2)


""" Rita Data Loader - Excel2Web -> loader danych z arkusza Excela do formularzy na stronie Web. """

numerPozycjiDict = {"1": "1. Koordynator projektu", "2": "2. Specjalista ds. finansów",
                    "3": "3. Koszty biura PL (czynsz, media, telefon, Internet, usługi pocztowe, koszty bankowe, mat. biurowe)",
                    "4": "4. Koordynator UA",
                    "5": "5. Specjalista ds. finansów UA",
                    "11": "11. podróże ekspertów PL (3 os. x 1000 zł)",
                    "12": "12. noclegi, wyżywienie i diety ekspertów PL (3 os. x 4 dni)",
                    "13": "13. catering na spotkania (spotkanie organizacyjne - 10 os. x 15 zł; warsztaty - 20 os. x 2 dni x 15 zł)",
                    "14": "14. wynajem sali na spotkania",
                    "15": "15. materiały biurowe na spotkania",
                    "16": "16. honoraria ekspertów PL (2 os. x 3 dni)",
                    "21": "21. przeprowadzenie ankiet i wywiadów",
                    "22": "22. zakup sprzętu dla uczelni (2 laptopy i drukarka)",
                    "23": "23. opracowanie wyników badania",
                    "24": "24. przygotowanie map / planów / makiety terenu na warsztaty",
                    "31": "31. podróże ekspertów PL (4 os. x 1000 zł)",
                    "32": "32. noclegi, wyżywienie i diety ekspertów PL (4 os. x 5 dni)",
                    "33": "33. catering na spotkania (spotkanie organizacyjne: 10 os. x 15 zł + warsztaty: 25 os. x 2 dni x 15 zł)",
                    "34": "34. honoraria ekspertów PL (3 os. x 4 dni x 500 zł)",
                    "35": "35. materiały biurowe na warsztaty",
                    "36": "36. wynajem sali na spotkania i warsztaty",
                    "41": "41. budżet na realizację projektu wypracowanego podczas warsztatów",
                    "51": "51. podróże ekspertów PL (3 os. x 1000 zł)",
                    "52": "52. noclegi, wyżywienie i diety ekspertów PL (3 os. x 3 dni)",
                    "53": "53. catering na spotkania (spotkanie ewaluacyjne 10 os. x 15 zl, spotkanie publiczne - prezentacja wyników 30 os. x 15 zl)",
                    "54": "54. wynajem sali na spotkania",
                    "55": "55. materiały biurowe na spotkania",
                    "56": "56. robocze tłumaczenia polsko-ukraińskie",
                    "57": "57. honoraria dla ekspertów "}

wierszDanych = []
szykDanych = [8, 9, 10, 11, 5, 7, 12, 13, 14, 15, 16]
tabelkaRita = openpyxl.load_workbook(filename="Rita_tabelka.xlsx", data_only=True)
arkusz = tabelkaRita.get_sheet_by_name("Wydatki 2015")

rita = Formularz()

licznik = 3
pozBud = arkusz.cell(row=licznik, column=5).value

while pozBud != None:
    pozBud = arkusz.cell(row=licznik, column=5).value
    if pozBud != None:
        for kolumna in szykDanych:
            tmp = re.sub('"', '', unicode(arkusz.cell(row=licznik, column=kolumna).value))
            if kolumna == 5:
                tmp = numerPozycjiDict[str(tmp)]
            if (kolumna == 10 or kolumna == 11) and tmp != "None":
                czas = datetime.datetime.strptime(tmp, '%Y-%m-%d %H:%M:%S')
                tmp = unicode(czas.strftime("%d.%m.%Y"))
            elif (kolumna == 10 or kolumna == 1) and tmp == "None":
                tmp = unicode(datetime.datetime.now().strftime("%d.%m.%Y"))
            if tmp == "None":
                tmp = " "
            wierszDanych.append(tmp)
        try:
            rita.dodaj(wierszDanych)
        except:
            print "Nie udało się dodać wiersza danych:"
            for kom in wierszDanych:
                print kom
        del wierszDanych[:]
    licznik += 1

rita.wyloguj()